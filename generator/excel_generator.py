from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import tempfile
import os
from collections import Counter


def generate_orders_excel(orders_data, output_path=None):

    # Генерирует Excel файл со всеми заказами и аналитикой
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"

        # Стили
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        company_font = Font(bold=True, size=16, color="2F5496")
        summary_font = Font(bold=True, size=14, color="2F5496")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Заголовок компании
        ws.merge_cells('A1:H1')
        ws['A1'] = 'ООО "Warehouse"'
        ws['A1'].font = company_font
        ws['A1'].alignment = center_align

        # Заголовки столбцов
        headers = [
            'Заказываемое количество',
            'Номер товара',
            'Описание',
            'Общая стоимость',
            'ФИО покупателя',
            'ФИО сотрудника',
            'Дата заказа',
            'Статус'
        ]

        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style

        # Заполняем данными заказов
        row_num = 4
        total_revenue = 0
        total_items = 0
        status_counter = Counter()
        buyer_counter = Counter()

        for order in orders_data:
            for item in order['items']:
                total_revenue += item['total_cost']
                total_items += item['quantity']
                status_counter[order['order_status']] += 1
                buyer_counter[order['buyer_name']] += 1

                # Заказываемое количество
                ws.cell(row=row_num, column=1, value=item['quantity']).alignment = center_align
                ws.cell(row=row_num, column=1).border = border_style

                # Номер товара
                ws.cell(row=row_num, column=2, value=item['product_id']).alignment = center_align
                ws.cell(row=row_num, column=2).border = border_style

                # Описание
                ws.cell(row=row_num, column=3, value=item['product_name']).alignment = left_align
                ws.cell(row=row_num, column=3).border = border_style

                # Общая стоимость
                ws.cell(row=row_num, column=4, value=f"{item['total_cost']:.2f} руб.").alignment = center_align
                ws.cell(row=row_num, column=4).border = border_style

                # ФИО покупателя
                ws.cell(row=row_num, column=5, value=order['buyer_name']).alignment = left_align
                ws.cell(row=row_num, column=5).border = border_style

                # ФИО сотрудника
                ws.cell(row=row_num, column=6, value=order['employee_name']).alignment = left_align
                ws.cell(row=row_num, column=6).border = border_style

                # Дата заказа
                ws.cell(row=row_num, column=7, value=order['order_date']).alignment = center_align
                ws.cell(row=row_num, column=7).border = border_style

                # Статус
                ws.cell(row=row_num, column=8, value=order['order_status']).alignment = center_align
                ws.cell(row=row_num, column=8).border = border_style

                row_num += 1

        # Итоговая строка для таблицы заказов
        ws.cell(row=row_num, column=3, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=row_num, column=3).alignment = right_align
        ws.cell(row=row_num, column=4, value=f"{total_revenue:.2f} руб.").font = Font(bold=True)
        ws.cell(row=row_num, column=4).alignment = center_align
        ws.cell(row=row_num, column=4).fill = PatternFill(start_color="FFFF00", fill_type="solid")

        # Добавляем аналитику ниже таблицы
        analytics_start_row = row_num + 3

        # Заголовок аналитики
        ws.merge_cells(f'A{analytics_start_row}:H{analytics_start_row}')
        ws[f'A{analytics_start_row}'] = '📊 СВОДНАЯ АНАЛИТИКА'
        ws[f'A{analytics_start_row}'].font = summary_font
        ws[f'A{analytics_start_row}'].alignment = center_align
        ws[f'A{analytics_start_row}'].fill = PatternFill(start_color="E6E6FA", fill_type="solid")

        # Основные показатели
        analytics_data = [
            ["Показатель", "Значение", "", "Статистика по статусам", "Количество"],
            ["Всего заказов", len(orders_data), "", "В обработке", status_counter.get("В обработке", 0)],
            ["Всего товаров", total_items, "", "Подтвержден", status_counter.get("Подтвержден", 0)],
            ["Общая выручка", f"{total_revenue:.2f} руб.", "", "В сборке", status_counter.get("В сборке", 0)],
            ["Средний чек", f"{total_revenue / max(len(orders_data), 1):.2f} руб.", "", "Готов к отгрузке",
             status_counter.get("Готов к отгрузке", 0)],
            ["Топ покупатель", get_top_buyer(buyer_counter), "", "Отгружен", status_counter.get("Отгружен", 0)],
            ["", "", "", "Доставлен", status_counter.get("Доставлен", 0)],
            ["", "", "", "Отменен", status_counter.get("Отменен", 0)]
        ]

        # Записываем аналитику
        for row_offset, row_data in enumerate(analytics_data):
            current_row = analytics_start_row + 1 + row_offset
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                if row_offset == 0:  # Заголовки
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F0F0F0", fill_type="solid")
                elif value and col in [2, 5]:  # Значения
                    cell.font = Font(bold=True)

        # Настраиваем ширину столбцов
        column_widths = {
            'A': 18,  # Заказываемое количество
            'B': 12,  # Номер товара
            'C': 30,  # Описание
            'D': 16,  # Общая стоимость
            'E': 25,  # ФИО покупателя
            'F': 25,  # ФИО сотрудника
            'G': 12,  # Дата заказа
            'H': 15  # Статус
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Автофильтр для заголовков
        ws.auto_filter.ref = f"A3:H{row_num - 1}"

        # Замораживаем заголовки
        ws.freeze_panes = "A4"

        # Создаем временный файл если путь не указан
        if not output_path:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                output_path = tmp_file.name

        # Сохраняем файл
        wb.save(output_path)
        return output_path

    except Exception as e:
        raise Exception(f"Ошибка при генерации Excel файла: {str(e)}")


def get_top_buyer(buyer_counter):
    # Возвращает топ покупателя
    if not buyer_counter:
        return "Нет данных"
    top_buyer = buyer_counter.most_common(1)[0]
    return f"{top_buyer[0]} ({top_buyer[1]} зак.)"


def prepare_all_orders_data(orders, db_session):

    # Подготавливает данные всех заказов для экспорта
    orders_data = []

    for order in orders:
        # Вычисляем общую стоимость заказа
        items_data = []

        for item in order.items:
            item_total = item.product.price * item.quantity

            items_data.append({
                'quantity': item.quantity,
                'product_id': item.product.id_Product,
                'product_name': item.product.Name_tov,
                'total_cost': item_total
            })

        order_data = {
            'order_code': order.Code,
            'buyer_name': order.buyer.Buyer_name if order.buyer else 'Не указан',
            'employee_name': order.employee.FIO if order.employee else 'Не указан',
            'order_date': datetime.now().strftime('%d.%m.%Y'),
            'order_status': order.status if order.status else 'В обработке',
            'items': items_data
        }

        orders_data.append(order_data)

    return orders_data


# Добавляем выравнивание по правому краю
right_align = Alignment(horizontal='right', vertical='center')